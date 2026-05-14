from __future__ import annotations

import csv
from datetime import datetime, timezone
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from gitbroom.core.models import BranchInfo


def _relative_time(dt: datetime) -> str:
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    delta = datetime.now(tz=timezone.utc) - dt
    days = delta.days
    if days == 0:
        hours = delta.seconds // 3600
        return f"{hours} saat önce" if hours > 0 else "az önce"
    if days < 30:
        return f"{days} gün önce"
    months = days // 30
    if months < 12:
        return f"{months} ay önce"
    return f"{days // 365} yıl önce"


def _location(b: BranchInfo) -> str:
    parts = []
    if b.is_local:
        parts.append("Local")
    if b.is_remote:
        parts.append("Remote")
    return " + ".join(parts)


def _rows(branches: list[BranchInfo]) -> list[dict]:
    return [
        {
            "Branch": b.name,
            "Yazar": b.last_commit_author,
            "Son Commit": _relative_time(b.last_commit_date),
            "Merge": f"{b.merge_type.value}" if b.is_merged else "Unmerged",
            "Risk": b.risk_score.label,
            "Konum": _location(b),
        }
        for b in branches
    ]


_HEADERS = ["Branch", "Yazar", "Son Commit", "Merge", "Risk", "Konum"]


class ExportManager:
    def to_csv(self, branches: list[BranchInfo], path: Path) -> None:
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=_HEADERS)
            writer.writeheader()
            writer.writerows(_rows(branches))

    def to_markdown(self, branches: list[BranchInfo], path: Path) -> None:
        lines = ["| " + " | ".join(_HEADERS) + " |"]
        lines.append("|" + "|".join("---" for _ in _HEADERS) + "|")
        for row in _rows(branches):
            lines.append("| " + " | ".join(row[h] for h in _HEADERS) + " |")
        path.write_text("\n".join(lines), encoding="utf-8")

    def to_excel(self, branches: list[BranchInfo], path: Path) -> None:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Branches"

        header_fill = PatternFill("solid", fgColor="2D3250")
        header_font = Font(bold=True, color="FFFFFF")
        risk_colors = {
            "green":  "A6E3A1",
            "yellow": "F9E2AF",
            "orange": "FAB387",
            "red":    "F38BA8",
        }

        for col, header in enumerate(_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, b in enumerate(branches, 2):
            data = [
                b.name,
                b.last_commit_author,
                _relative_time(b.last_commit_date),
                b.merge_type.value if b.is_merged else "Unmerged",
                b.risk_score.label,
                _location(b),
            ]
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 5:  # Risk column
                    color = risk_colors.get(b.risk_score.level.value, "FFFFFF")
                    cell.fill = PatternFill("solid", fgColor=color)

        # Auto-fit column widths
        col_widths = [60, 25, 15, 15, 18, 18]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

        ws.freeze_panes = "A2"
        wb.save(path)

    def to_pdf(self, branches: list[BranchInfo], path: Path, repo_name: str = "") -> None:
        from PyQt6.QtGui import QTextDocument
        from PyQt6.QtPrintSupport import QPrinter

        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
        printer.setOutputFileName(str(path))
        printer.setPageMargins(
            printer.pageLayout().margins()
            .__class__(15, 15, 15, 15)
        )

        rows = _rows(branches)
        color_map = {
            "green":  "#a6e3a1",
            "yellow": "#f9e2af",
            "orange": "#fab387",
            "red":    "#f38ba8",
        }

        header_cells = "".join(
            f"<th style='background:#2d3250;color:#fff;padding:6px 10px'>{h}</th>"
            for h in _HEADERS
        )
        row_cells = []
        for b, row in zip(branches, rows):
            bg = color_map.get(b.risk_score.level.value, "#ffffff")
            cells = []
            for h in _HEADERS:
                cell_bg = bg if h == "Risk" else "transparent"
                cells.append(
                    f"<td style='padding:5px 10px;background:{cell_bg}'>{row[h]}</td>"
                )
            row_cells.append(f"<tr>{''.join(cells)}</tr>")

        html = f"""
        <html><body style='font-family:Arial,sans-serif;font-size:11px'>
        <h2>GitBroom — Branch Raporu{': ' + repo_name if repo_name else ''}</h2>
        <p style='color:#888'>{datetime.now().strftime('%d.%m.%Y %H:%M')} · {len(branches)} branch</p>
        <table border='0' cellspacing='0' cellpadding='0'
               style='border-collapse:collapse;width:100%'>
          <thead><tr>{header_cells}</tr></thead>
          <tbody>{''.join(row_cells)}</tbody>
        </table>
        </body></html>
        """

        doc = QTextDocument()
        doc.setHtml(html)
        doc.print(printer)
